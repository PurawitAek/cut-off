"""
Generate default_input.xlsx — the workbook the app loads as its default template.

Sheets
------
1. applicants   – 1 000 sample rows (customerID, product, segment, score, grade)
2. PD_table     – grade → PD (%) per the model defaults
3. E31_table    – grade → %Ever31@MOB3 per the model defaults
4. Economics    – per-product EIR / COF / OPEX / LGD defaults
5. AQI          – AQI parameter defaults (single reference row)

Run:
    python make_default_xlsx.py
Outputs:
    default_input.xlsx   (same directory)
"""
from __future__ import annotations
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from data import make_sample_data
from config import PD_DEFAULT, E31_DEFAULT, ECON_DEFAULT, AQI_DEFAULT, GRADE_BANDS

# ── colour palette ─────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="0E7C86")   # teal
ALT_FILL     = PatternFill("solid", fgColor="EAF4F5")   # light teal
GOOD_FILL    = PatternFill("solid", fgColor="D6EFE5")   # soft green
MID_FILL     = PatternFill("solid", fgColor="FDF3E0")   # soft amber
BAD_FILL     = PatternFill("solid", fgColor="FDECEA")   # soft red
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
TITLE_FONT   = Font(name="Calibri", bold=True, size=12, color="0E7C86")
THIN = Side(border_style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws, row: int, cols: list[str]):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _data_row(ws, row: int, values: list, fill=None):
    f = fill if fill else (ALT_FILL if row % 2 == 0 else WHITE_FILL)
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = BODY_FONT
        cell.fill = f
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")


def _auto_width(ws, extra: int = 2):
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col_cells if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + extra, 40)


# ── Sheet 1 : applicants ──────────────────────────────────────────────────────

def _sheet_applicants(wb):
    ws = wb.active
    ws.title = "applicants"

    ws["A1"] = "Credit Decisioning Workbench — Sample Applicant Data"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = "Upload this sheet (or replace with your own data) via the sidebar file uploader."
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="5C6B7A")
    ws.merge_cells("A2:E2")

    cols = ["id", "product", "segment", "score", "grade"]
    _header_row(ws, 3, cols)
    ws.row_dimensions[3].height = 22

    df = make_sample_data(n=1000, seed=42)

    grade_fills = {
        1: GOOD_FILL, 2: GOOD_FILL, 3: GOOD_FILL,
        4: MID_FILL,  5: MID_FILL,  6: MID_FILL,
        7: BAD_FILL,  8: BAD_FILL,  9: BAD_FILL, 10: BAD_FILL,
    }

    for i, row in df.iterrows():
        r = i + 4
        g = int(row["grade"])
        values = [row["id"], row["product"], row["segment"],
                  int(row["score"]), g]
        _data_row(ws, r, values, fill=grade_fills.get(g))

    # freeze header rows
    ws.freeze_panes = "A4"
    _auto_width(ws)

    # score column: number format
    for r in range(4, 4 + len(df)):
        ws.cell(row=r, column=4).number_format = "0"
        ws.cell(row=r, column=5).number_format = "0"


# ── Sheet 2 : PD_table ────────────────────────────────────────────────────────

def _sheet_pd(wb):
    ws = wb.create_sheet("PD_table")
    ws["A1"] = "Grade → Probability of Default (%)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws["A2"] = "Edit PD values in the sidebar 'Grade → PD (%)' expander, or use this table as reference."
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="5C6B7A")
    ws.merge_cells("A2:B2")

    _header_row(ws, 3, ["Grade", "PD (%)"])
    for i, (g, pd) in enumerate(zip(GRADE_BANDS, PD_DEFAULT), 4):
        grade_fill = GOOD_FILL if g <= 3 else (MID_FILL if g <= 6 else BAD_FILL)
        ws.cell(row=i, column=1, value=g).font = BODY_FONT
        ws.cell(row=i, column=1).fill = grade_fill
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=round(pd, 4)).font = BODY_FONT
        ws.cell(row=i, column=2).fill = grade_fill
        ws.cell(row=i, column=2).border = BORDER
        ws.cell(row=i, column=2).number_format = "0.0000"
    _auto_width(ws)


# ── Sheet 3 : E31_table ───────────────────────────────────────────────────────

def _sheet_e31(wb):
    ws = wb.create_sheet("E31_table")
    ws["A1"] = "Grade → %Ever31@MOB3  (Path-3 AQI input)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws["A2"] = "Used by the AQI chain to compute blended Ever31@MOB3 for the approved portfolio."
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="5C6B7A")
    ws.merge_cells("A2:B2")

    _header_row(ws, 3, ["Grade", "Ever31@MOB3 (%)"])
    for i, (g, e31) in enumerate(zip(GRADE_BANDS, E31_DEFAULT), 4):
        grade_fill = GOOD_FILL if g <= 3 else (MID_FILL if g <= 6 else BAD_FILL)
        ws.cell(row=i, column=1, value=g).font = BODY_FONT
        ws.cell(row=i, column=1).fill = grade_fill
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2, value=round(e31, 4)).font = BODY_FONT
        ws.cell(row=i, column=2).fill = grade_fill
        ws.cell(row=i, column=2).border = BORDER
        ws.cell(row=i, column=2).number_format = "0.0000"
    _auto_width(ws)


# ── Sheet 4 : Economics ───────────────────────────────────────────────────────

def _sheet_econ(wb):
    ws = wb.create_sheet("Economics")
    ws["A1"] = "Per-product economics defaults"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = "Paste your own product rows; adjust in the sidebar 'Economics per product' expander."
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="5C6B7A")
    ws.merge_cells("A2:F2")

    cols = ["product", "Avg Loan (THB)", "EIR (%)", "COF (%)", "OPEX / CAC (THB)", "LGD (%)"]
    _header_row(ws, 3, cols)

    for i, (prod, v) in enumerate(ECON_DEFAULT.items(), 4):
        f = ALT_FILL if i % 2 == 0 else WHITE_FILL
        values = [prod, v["loan"], round(v["eir"]*100, 4), round(v["cof"]*100, 4),
                  v["opex"], round(v["lgd"]*100, 2)]
        _data_row(ws, i, values, fill=f)
        ws.cell(row=i, column=2).number_format = "#,##0"
        ws.cell(row=i, column=3).number_format = "0.00"
        ws.cell(row=i, column=4).number_format = "0.00"
        ws.cell(row=i, column=5).number_format = "#,##0"
        ws.cell(row=i, column=6).number_format = "0.0"
    _auto_width(ws)


# ── Sheet 5 : AQI ─────────────────────────────────────────────────────────────

def _sheet_aqi(wb):
    ws = wb.create_sheet("AQI")
    ws["A1"] = "AQI parameter defaults"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws["A2"] = "Configure these in the sidebar 'AQI parameters' expander."
    ws["A2"].font = Font(name="Calibri", italic=True, size=9, color="5C6B7A")
    ws.merge_cells("A2:B2")

    _header_row(ws, 3, ["Parameter", "Default value"])
    rows = [
        ("% Avg Credit Cost / yr  (cc)",  AQI_DEFAULT["cc"]),
        ("LGD %                   (lgd)", AQI_DEFAULT["lgd"]),
        ("PD roll 31→91 %         (pd)",  AQI_DEFAULT["pd"]),
        ("Loss-curve factor       (lc)",   AQI_DEFAULT["lc"]),
    ]
    for i, (label, val) in enumerate(rows, 4):
        f = ALT_FILL if i % 2 == 0 else WHITE_FILL
        ws.cell(row=i, column=1, value=label).font = BODY_FONT
        ws.cell(row=i, column=1).fill = f
        ws.cell(row=i, column=1).border = BORDER
        ws.cell(row=i, column=2, value=val).font = BODY_FONT
        ws.cell(row=i, column=2).fill = f
        ws.cell(row=i, column=2).border = BORDER
        ws.cell(row=i, column=2).number_format = "0.0000"
    _auto_width(ws)


# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "default_input.xlsx")
    wb = openpyxl.Workbook()
    _sheet_applicants(wb)
    _sheet_pd(wb)
    _sheet_e31(wb)
    _sheet_econ(wb)
    _sheet_aqi(wb)
    wb.save(out)
    print(f"Saved → {out}")
    print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
